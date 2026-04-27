import hashlib
from pathlib import Path
import openpyxl
from app.core.config import settings


def compute_master_sheet_sha256(file_path: str) -> str:
    """
    Compute SHA-256 hash of MASTER sheet content only.
    Ignores Excel metadata, macros, and formatting.
    Only cell values are hashed — this is what matters for deduplication.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=False, data_only=True)

    if settings.SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"{settings.SHEET_NAME} sheet not found in {Path(file_path).name}")

    ws = wb[settings.SHEET_NAME]
    sha256 = hashlib.sha256()

    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                sha256.update(str(cell).encode("utf-8"))

    wb.close()
    return sha256.hexdigest()


def compute_sha256(file_path: str) -> str:
    """SHA-256 of MASTER sheet cell values (same as compute_master_sheet_sha256)."""
    return compute_master_sheet_sha256(file_path)